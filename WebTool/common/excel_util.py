import openpyxl
import os
import json
from collections import namedtuple




def create_xlsHeader():
    return  namedtuple('xlsHeader', ['strType', 'strVarName', 'bVector''strClassMemberName'])

def init_json_string(bToMap,strSaveName):
    strJson =''
    if bToMap == True :
        strJson += '{\n'
        strJson += '\n \"strType\":\"'
        strJson += strSaveName
        strJson += '\",\n'
        strJson += '\"m_mapTableData\":{\n'
    else :
        strJson += '{\n'
        strJson += '\n \"strType\:\"'
        strJson += strSaveName
        strJson += '\",\n'
        strJson += '\"vec'
        strJson += strSaveName
        strJson += '\":\n [\"'
    return strJson

def GetMemberName( strTypeName,strMemberName):
    if strTypeName == 'bool' :
        return 'm_b' + strMemberName
    elif strTypeName == 'tinyint' or strTypeName == 'char' :
        return 'm_t' + strMemberName
    elif strTypeName == 'uchar' :
        return 'm_ut' + strMemberName
    elif strTypeName == 'short' :
        return 'm_n' + strMemberName
    elif strTypeName == 'ushort' :
        return 'm_un' + strMemberName
    elif strTypeName == 'int' :
        return 'm_i' + strMemberName
    elif strTypeName == 'uint' :
        return 'm_ui' + strMemberName
    elif strTypeName == 'bint' :
        return 'm_bi' + strMemberName
    elif strTypeName == 'String' or strTypeName == 'string' :
        return 'm_str' + strMemberName
    elif strTypeName == 'float' :
        return 'm_f' + strMemberName
    else :
        return ''

def type_check( value ) :
    try:
        cValue = float(value)
    except Exception as e :
        return False
    return True

def GetXlsHeader(lstHeader,exSheet) :
    nRet = 0
    for i in range(1, exSheet.max_row + 1, 1):
        nRet = i
        if exSheet.cell(row=i, column=1).value == '#' :
            continue;
        else :
            break;

    cTypeRow = list(exSheet.rows)[nRet -1]
    nRet += 1
    cVarRow = list(exSheet.rows)[nRet -1];
    nRet += 1

    for i in range(0, exSheet.max_column, 1):
        cHeader = create_xlsHeader()
        cHeader.strVarName =  cVarRow[i].value
        strValue = cTypeRow[i].value
        if strValue[0] =='*' :
            strValue = strValue[1:]
        cHeader.strType = strValue
        strVector = cHeader.strVarName.split('_')
        cHeader.bVector = len(strVector) > 1
        cHeader.strClassMemberName = ""
        str1 =''
        if cHeader.bVector == True :
            check_value = strVector[len(strVector)-1]
            isNumber = type_check(check_value)

            if isNumber == True :
                for idx in range(0,len(strVector) -1,1) :
                    str1 += strVector[idx];
                cHeader.strVarName = str1;
            else :
                for idx in range(0, len(strVector), 1):
                    str1 += strVector[idx]
                    if idx < len(strVector) -1 :
                        cHeader.strClassMemberName += strVector[idx]
        lstHeader.append(cHeader)
    return nRet

def GetExcelTypeValue(strTargetValue, strTypeName):
    if strTypeName == 'string' or strTypeName == 'String' :
        strValue = '\"'
        strValue += str(strTargetValue)
        strValue += '\"'
        return strValue
    return str(strTargetValue)


def excel_to_json_new( strfile ,bToMap = True):
    excel_document = openpyxl.load_workbook(strfile)
    #print(excel_document.get_sheet_names())
    file_name = os.path.basename(strfile)
    #print( file_name)
    sheet_name = file_name.split('.')[0]
    #print( sheet_name )
    cSheet = excel_document.get_sheet_by_name(sheet_name)

    strJson = init_json_string(True,sheet_name)
    #print( cSheet.max_column , cSheet.max_row)

    nVectorCount = 0
    nVectorItemCount = 0
    nVectorItemLoop = 0
    lstHeader =[]

    # 헤더 추출
    nRow = GetXlsHeader(lstHeader, cSheet)

    for irow in range(nRow, cSheet.max_row+1, 1):
        if cSheet.cell(row=irow, column=1).value == '#' :
            continue;

        if False == bToMap :
            strJson += '{'
        strFirstValue =''

        for i in range(0, len(lstHeader), 1):
            cHaeder = lstHeader[i]
            if cHaeder.bVector == True and nVectorCount == 0 :
                strJson += '\n'
                if cHaeder.strClassMemberName == '' :
                    strJson += '\"'
                    strJson += cHaeder.strVarName
                    strJson += '\":['
                    nVectorItemCount = 1
                else :
                    strJson += '\"'
                    strJson += cHaeder.strClassMemberName
                    strJson += '\":['
                    indexs = [i for i, x in enumerate(lstHeader) if x.strClassMemberName == cHaeder.strClassMemberName]
                    nVectorItemCount = len(indexs)

                indexs = [i for i, x in enumerate(lstHeader) if x.strVarName == cHaeder.strVarName]
                nVectorCount = len(indexs)
                strJson += '\n'

            if nVectorCount > 0 :
                if nVectorItemLoop == 0 and cHaeder.strClassMemberName != "":
                    strJson += '{\n'

                strVecValue = cSheet.cell(row=irow, column=i +1).value

                if cHaeder.strClassMemberName == "" :
                    strJson += GetExcelTypeValue(strVecValue, cHaeder.strType)
                    strJson += ','
                else :
                    strJson += '\"'
                    strJson += GetMemberName(cHaeder.strType, cHaeder.strVarName)
                    strJson += '\":'
                    strJson += GetExcelTypeValue(strVecValue, cHaeder.strType)
                    strJson += ','

                nVectorItemLoop += 1
                if nVectorItemLoop == nVectorItemCount :
                    nVectorItemLoop = 0
                    nVectorCount -= 1
                    if cHaeder.strClassMemberName != "" :
                        strJson = strJson[:-1]  # 마지낙 컴마 제거
                        strJson += '},'

                if nVectorCount <= 0 :
                    strJson = strJson[:-1];
                    strJson += "],";
                continue

            strTValue = cSheet.cell(row=irow, column=i +1).value

            if True == bToMap and strFirstValue == '' :
                strFirstValue = strTValue
                strJson += '\"'
                strJson += str(strFirstValue)
                strJson += '\":'
                strJson += '{'

            strJson += '\n'
            strJson += '\"'
            strJson += GetMemberName(cHaeder.strType, cHaeder.strVarName)
            strJson += '\":'
            strJson += GetExcelTypeValue(strTValue, cHaeder.strType)
            strJson += ','

        strJson = strJson[:-1]
        strJson += '},'

    strJson = strJson[:-1]
    if bToMap == True :
        strJson += '}\n}'
    else :
        strJson += ']\n}'

    return sheet_name ,strJson


def excel_to_json( strfile ,bToMap = True):

    excel_document = openpyxl.load_workbook(strfile)
    #print(excel_document.get_sheet_names())
    file_name = os.path.basename(strfile)
    #print( file_name)
    sheet_name = file_name.split('.')[0]
    #print( sheet_name )
    cSheet = excel_document.get_sheet_by_name(sheet_name)

    strJson = init_json_string(True,sheet_name)
    lstType =[]
    lstVarNames =[]
    #print( cSheet.max_column , cSheet.max_row)

    nVectorCount = 0
    nVectorItemCount = 0
    nVectorItemLoop = 0

    for irow in range(1, cSheet.max_row+1, 1):
        if cSheet.cell(row=irow, column=1).value == '#' :
            continue;
        #헤더 추출
        if len(lstType) == 0 :
            for icol in range(1, cSheet.max_column +1, 1):
                strType = cSheet.cell(row=irow,column=icol).value
                lstType.append(strType)
            continue;

        if len(lstVarNames) == 0 :
            for icol in range(1, cSheet.max_column +1, 1):
                strVarName = cSheet.cell(row=irow,column=icol).value
                lstVarNames.append(strVarName)
            continue;
        #print(cSheet.cell(row=irow,column=2).value)
        if False == bToMap :
            strJson += '{'
        strFirstValue =''

        for idx in range(0, len(lstVarNames), 1):
            strVector = lstVarNames[idx].split('_')
            if strVector[0] == 'c' :
                nVectorItemCount = strVector[1]
                strName = strVector[2]
                nVectorCount = strVector[3]
                strJson += '\n'
                strJson += '\"c'
                strJson += strName
                strJson += '\":['
                continue

            if nVectorCount > 0 :
                if nVectorItemLoop == 0 :
                    strJson += '{\n'

                strVecValue = cSheet.cell(row=irow, column=idx +1).value
                strJson += '\"'
                strJson += GetMemberName(lstType[idx], lstVarNames[idx])
                strJson +='\":'
                strJson += GetExcelTypeValue(strVecValue, lstType[idx])
                strJson += ','

                nVectorItemLoop += 1

                if nVectorItemLoop == nVectorItemCount :
                    nVectorItemLoop = 0
                    nVectorCount -= 1
                    strJson = strJson[:-1] #마지낙 컴마 제거
                    strJson += '},'

                if nVectorCount <= 0 :
                    strJson = strJson[:-1]
                    strJson += ']'
                continue

            strTValue = cSheet.cell(row=irow, column=idx +1).value

            if True == bToMap and strFirstValue == '' :
                strFirstValue = strTValue
                strJson += '\"'
                strJson += str(strFirstValue)
                strJson += '\":'
                strJson += '{'

            strJson += '\n'
            strJson += '\"'
            strJson +=  GetMemberName(lstType[idx], lstVarNames[idx])
            strJson += '\":'
            strJson += GetExcelTypeValue(strTValue, lstType[idx])
            strJson += ','

        strJson = strJson[:-1]
        strJson += '},'

    strJson = strJson[:-1]
    if bToMap == True :
        strJson += '}\n}'
    else :
        strJson += ']\n}'

    return sheet_name ,strJson

def find_all_excel_file(strdir):
    #lstFiles =[]
    #for root, dirs, files in os.walk(strdir):
    #    for filename in files:
    #        lstFiles.append((filename))
    #print(lstFiles)
    included_extenstions = ['.xls', '.xlsm', '.json']
    file_names = [fn for fn in os.listdir(strdir)
                  if any(fn.endswith(ext) for ext in included_extenstions)]
    #print(file_names)
    return file_names


def json_table_file_read(strfile):
    file_name = os.path.basename(strfile)
    save_name = file_name.split('.')[0]

    fd = open(strfile,'r',encoding='UTF-8')
    lines = fd.readlines()
    strJsonTxt = '{\"strType\":\"'
    strJsonTxt += save_name
    strJsonTxt += '\",\n'
    strJsonTxt += '\"m_mapTableData\":'
    for line in lines:
        #print(line)
        strJsonTxt += line.split('//')[0]
    strJsonTxt += '}'
    fd.close()

    return strJsonTxt